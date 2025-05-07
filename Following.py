#!/usr/bin/env python
# coding: utf-8

# In[16]:


# Shortlist Follower Chart

import pandas as pd
from scipy.stats import linregress
from tikapi import TikAPI, ValidationException, ResponseException
from openpyxl import Workbook, load_workbook
from datetime import datetime, timedelta

# Initialize TikAPI
api = TikAPI("8mqoTQs1AXfSs6nskRCr5obvsWVytvQ1J0YPvIS1ylfEtl2D")
excel_file = "/Users/jw/Downloads/shortlist_data.xlsx"

shortlist = []


try:
    # Fetch the initial response
    response = api.public.followingList(
        secUid="MS4wLjABAAAAboanSl94WMrjvJtHejLumdRGgy9oYuygOQfbC-iVne34BIfjcygpqSH84qsh2XcT"
    )

    # Prepare the Excel file
    try:
        workbook = load_workbook(excel_file)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Username", "Date", "Follower Count"])

    while response:
        # Extract relevant data
        data = response.json()
        user_list = data.get("userList", [])
        current_date = datetime.now().strftime("%Y-%m-%d")

        
        for user_data in user_list:
            username = user_data['user']['uniqueId']
            shortlist.append(username)
            follower_count = user_data['stats']['followerCount']
            sheet.append([username, current_date, follower_count])

        workbook.save(excel_file)

        # Check for the next cursor
        next_cursor = data.get("nextCursor")
        if not next_cursor:
            break

        print(f"Getting next items {next_cursor}")
        response = response.next_items()

    print(shortlist)

except ValidationException as e:
    print(e, e.field)

except ResponseException as e:
    print(e, e.response.status_code)

today = datetime.today().strftime('%m.%d.%y')

# Load the data
file_path = '/Users/jw/Downloads/shortlist_data.xlsx'  # Replace with your Excel file path
data = pd.read_excel(file_path)

# Sort the data to ensure it is ordered by Username and Date
data = data.sort_values(by=['Username', 'Date'])

# Calculate the Daily Difference
data['Daily Difference'] = data.groupby('Username')['Follower Count'].diff().fillna(0)

# Save the updated DataFrame to a new Excel file
output_file = f'/Users/jw/Downloads/shortlist_data_updated{today}.xlsx'
data.to_excel(output_file, index=False)

print(f"Updated file saved to: {output_file}")

today = datetime.today().strftime('%m.%d.%y')

def calculate_slopes_with_current_followers(input_file, output_file):
    # Load the data
    excel_data = pd.ExcelFile(input_file)
    sheet_name = excel_data.sheet_names[0]  # Assumes data is in the first sheet
    data = excel_data.parse(sheet_name)

    # Ensure 'Date' is in datetime format
    data['Date'] = pd.to_datetime(data['Date'])

    # Initialize a list to store results
    results = []

    # Group data by 'Username' and calculate the slope, current follower count, and average percent change
    for username, group in data.groupby('Username'):
        # Sort data by date to ensure proper order
        group = group.sort_values(by='Date')
        
        # Convert dates to ordinal numbers (x-axis)
        x = group['Date'].map(lambda date: date.toordinal())
        y = group['Follower Count']
        
        # Calculate slope if sufficient data points exist
        if len(group) > 1:
            slope, _, _, _, _ = linregress(x, y)
        else:
            slope = None  # Not enough data points
        
        # Calculate average percent change, excluding the first entry and handling invalid values
        pct_changes = group['Follower Count'].pct_change().dropna()
        valid_pct_changes = pct_changes[~pct_changes.isin([float('inf'), float('-inf')])]  # Exclude #DIV/0!
        avg_pct_change = valid_pct_changes.mean() if not valid_pct_changes.empty else None
        
        # Append results for each username
        results.append({
            'Username': username, 
            'Slope': slope, 
            'Average Percent Change': avg_pct_change
        })

    # Convert results to a DataFrame
    slope_data = pd.DataFrame(results)

    # Save the output to a new sheet in the same Excel file
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write the original data
        data.to_excel(writer, index=False, sheet_name=sheet_name)
        # Write the slope results with average percent change
        slope_data.to_excel(writer, index=False, sheet_name='Slopes')

# Example usage
input_file = f'/Users/jw/Downloads/shortlist_data_updated{today}.xlsx'
output_file = f'/Users/jw/Downloads/shortlist_data_slope{today}.xlsx'
calculate_slopes_with_current_followers(input_file, output_file)

print('Done!')


# In[ ]:




